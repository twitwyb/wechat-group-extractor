# -*- coding: utf-8 -*-
"""
微信成员提取工具 - 导出模块
支持导出为Excel和CSV格式
"""

import os
import csv
from datetime import datetime
from typing import List, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    Workbook = None

from config import OUTPUT_DIR, EXCEL_FILENAME_PREFIX, CSV_FILENAME_PREFIX

# 列名映射: 内部字段 -> 中文表头
COLUMN_mapping = {
    "index": "序号",
    "wxid": "wxid",
    "nickname": "昵称",
    "remark": "备注名",
    "alias": "微信号",
}
EXPORT_COLUMNS = list(COLUMN_mapping.values())


def _generate_path(prefix: str, group_name: str, ext: str) -> str:
    """生成带时间戳的输出文件路径"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = "".join(c for c in group_name if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"{prefix}{safe_name}_{timestamp}.{ext}"
    return os.path.join(OUTPUT_DIR, filename)


def export_to_excel(
    members: List[Dict[str, str]],
    group_name: str,
    output_path: str = None
) -> Optional[str]:
    """导出为Excel格式"""
    if Workbook is None:
        print("[错误] 未安装openpyxl，请运行: pip install openpyxl")
        return None

    try:
        if not output_path:
            output_path = _generate_path(EXCEL_FILENAME_PREFIX, group_name, "xlsx")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "群成员"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        fields = ["index", "wxid", "nickname", "remark", "alias"]
        for row_idx, member in enumerate(members, 2):
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=row_idx, column=col_idx, value=member.get(field, ""))

        # 列宽
        for col, width in {"A": 8, "B": 25, "C": 20, "D": 20, "E": 20}.items():
            ws.column_dimensions[col].width = width

        # 边框
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        for row in ws.iter_rows(min_row=1, max_row=len(members) + 1, max_col=len(EXPORT_COLUMNS)):
            for cell in row:
                cell.border = thin_border

        ws.freeze_panes = "A2"
        wb.save(output_path)
        print(f"[+] 已导出Excel: {output_path}")
        return output_path

    except Exception as e:
        print(f"[错误] 导出Excel失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def export_to_csv(
    members: List[Dict[str, str]],
    group_name: str,
    output_path: str = None
) -> Optional[str]:
    """导出为CSV格式"""
    try:
        if not output_path:
            output_path = _generate_path(CSV_FILENAME_PREFIX, group_name, "csv")

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(EXPORT_COLUMNS)
            for member in members:
                writer.writerow([
                    member.get("index", ""),
                    member.get("wxid", ""),
                    member.get("nickname", ""),
                    member.get("remark", ""),
                    member.get("alias", ""),
                ])

        print(f"[+] 已导出CSV: {output_path}")
        return output_path

    except Exception as e:
        print(f"[错误] 导出CSV失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def _print_summary(members: List[Dict[str, str]]) -> None:
    """打印导出摘要"""
    total = len(members)
    if total == 0:
        return
    with_wxid = sum(1 for m in members if m.get("wxid"))
    with_nickname = sum(1 for m in members if m.get("nickname"))
    with_alias = sum(1 for m in members if m.get("alias"))
    print(f"\n=== 导出摘要 ===")
    print(f"总成员数: {total}")
    print(f"有wxid: {with_wxid} ({with_wxid/total*100:.1f}%)")
    print(f"有昵称: {with_nickname} ({with_nickname/total*100:.1f}%)")
    print(f"有微信号: {with_alias} ({with_alias/total*100:.1f}%)")
    print(f"================\n")


def export_members(
    members: List[Dict[str, str]],
    group_name: str,
    export_format: str = "excel"
) -> Optional[str]:
    """
    导出成员数据的主函数
    参数:
        members: 成员信息列表
        group_name: 群名称
        export_format: 导出格式 ("excel" 或 "csv")
    返回: 输出文件路径或None
    """
    _print_summary(members)

    if export_format.lower() == "excel":
        return export_to_excel(members, group_name)
    elif export_format.lower() == "csv":
        return export_to_csv(members, group_name)
    else:
        print(f"[错误] 不支持的导出格式: {export_format}")
        return None
